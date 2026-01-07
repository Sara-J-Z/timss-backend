import os
import re
import threading
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from main_app.services.graph_upload_session import GraphUploadSessionClient

EXCEL_DIR = os.path.join(os.getcwd(), "excel_files")
os.makedirs(EXCEL_DIR, exist_ok=True)

# Lock per school to avoid concurrent local save/upload for same workbook
_SCHOOL_LOCKS: dict[str, threading.Lock] = {}


def safe_name(name: str) -> str:
    bad = ['/', '\\', ':', '*', '?', '"', '<', '>', '|']
    for ch in bad:
        name = name.replace(ch, '-')
    return name.strip()[:120] or "UnknownSchool"


def safe_sheet_name(name: str) -> str:
    r"""
    Excel sheet rules:
    - max length 31
    - cannot contain: \ / ? * [ ]
    """
    name = (name or "Sheet1").strip()
    name = re.sub(r"\s+", "_", name)          # spaces -> _
    name = re.sub(r"[\\/*?:\[\]]", "-", name) # forbidden -> -
    return name[:31] or "Sheet1"


def _get_lock_for_school(safe_school: str) -> threading.Lock:
    if safe_school not in _SCHOOL_LOCKS:
        _SCHOOL_LOCKS[safe_school] = threading.Lock()
    return _SCHOOL_LOCKS[safe_school]


def save_to_excel(data: dict):
    school_name = data.get("school_name", "UnknownSchool")
    subject_raw = data.get("subject", "UnknownSubject")
    subject = safe_sheet_name(subject_raw)

    safe_school = safe_name(school_name)

    remote_folder = safe_school
    remote_filename = f"{safe_school}.xlsx"
    file_path = os.path.join(EXCEL_DIR, remote_filename)

    lock = _get_lock_for_school(safe_school)

    with lock:
        client = GraphUploadSessionClient()

        # ✅ After deploy or after cleanup, local file may not exist.
        # Download the existing OneDrive workbook first to avoid overwriting history.
        if not os.path.exists(file_path):
            try:
                downloaded = client.download_file(remote_folder, remote_filename, file_path)
                if downloaded:
                    print("[OneDrive] Downloaded existing workbook before updating.")
                else:
                    print("[OneDrive] Workbook not found on OneDrive yet. Will create a new one locally.")
            except Exception as e:
                print(f"[OneDrive Download Error] {e}")
                # continue; we may create a new workbook locally if download fails

        # Load or create workbook
        if os.path.exists(file_path):
            wb = openpyxl.load_workbook(file_path)
        else:
            wb = openpyxl.Workbook()
            default_sheet = wb.active
            wb.remove(default_sheet)

        # Load or create sheet
        if subject in wb.sheetnames:
            ws = wb[subject]
        else:
            ws = wb.create_sheet(title=subject)

            base_headers = [
                "date", "time", "student_name", "class_name", "teacher_name",
                "school_operation_region", "auto_correct_score_points"
            ]
            answers = data.get("answers", [])
            question_headers = [ans.get("question_number") for ans in answers]
            ws.append(base_headers + question_headers)

            # Header styling
            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Append row
        row = [
            data.get("date"),
            data.get("time"),
            data.get("student_name"),
            data.get("class_name"),
            data.get("teacher_name"),
            data.get("school_operation_region"),
            data.get("auto_correct_score_points"),
        ]
        question_values = [ans.get("answer_value") for ans in data.get("answers", [])]
        ws.append(row + question_values)

        # Borders + zebra
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        for idx, row_cells in enumerate(ws.iter_rows(), 1):
            if idx != 1:
                fill_color = "DCE6F1" if idx % 2 == 0 else "FFFFFF"
            else:
                fill_color = None

            for cell in row_cells:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if fill_color:
                    cell.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        # Column widths
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 25

        # Save once
        wb.save(file_path)

        # Ensure file flushed to disk before upload
        try:
            with open(file_path, "rb") as f:
                os.fsync(f.fileno())
        except Exception:
            pass

        # Upload to OneDrive (replace)
        try:
            client.upload_large_file(
                local_path=file_path,
                remote_folder=remote_folder,
                remote_filename=remote_filename,
                chunk_size_mb=10,
                max_retries=1
            )

            print("[OneDrive Upload] Success. Cleaning local file...")

            # ✅ Delete local file after successful upload
            try:
                os.remove(file_path)
                print(f"[CLEANUP] Local file deleted: {file_path}")
            except Exception as e:
                print(f"[CLEANUP WARNING] Could not delete local file: {e}")

        except Exception as e:
            msg = str(e)

            # If locked, skip upload and keep local file for next attempt
            if "423" in msg or "Locked" in msg:
                print("[OneDrive Upload] Skipped (Locked). Will upload on next submission.")
            else:
                print(f"[OneDrive Upload Error] {e}")

    # NOTE: file may be deleted if upload succeeded
    return file_path
